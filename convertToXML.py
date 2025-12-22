#Convert Excel to XML
from io import BytesIO
import openpyxl
import re
import streamlit as st
import xml.etree.ElementTree as ET

st.set_page_config(layout="wide")

# ======================
# UI Header
# ======================
st.title("🧾 Konversi Excel to XML")

st.markdown("""
Alat bantu konversi **Template Bukti Potong Excel** menjadi **XML** yang dapat diimpor ke Coretax.
Dapat digunakan apabila tidak mempunyai microsoft excel untuk menyimpan file menjadi xml.

### Langkah Penggunaan:
1️⃣ Pilih Jenis SPT (SPT 21 / SPT Unifikasi)  
2️⃣ Upload file excel dari format Excel
4️⃣ Tekan **Proses** untuk menghasilkan file template
""")

# ======================
# STATE HANDLING
# ======================
if "step" not in st.session_state:
    st.session_state.step = "pilihSPT"

def go_to_step(step):
    st.session_state.step = step

# ======================
# Step 1: Pilih SPT
# ======================
if st.session_state.step == "pilihSPT":
    st.write("### 🧾 Langkah 1 — Pilih SPT yang ingin diubah")
    type_map = {
        "SPT Masa PPh 21": "Bp21", "SPT Masa Unifikasi": "Bpu"
    }

    type_spt_name = st.selectbox("🗓️ Pilih SPT:", list(type_map.keys()))
    type_spt = type_map[type_spt_name]

    st.session_state.type_spt = type_spt
    
    if st.button("📊 Lanjutkan"):
        go_to_step("upload")
        st.rerun()

# ======================
# Step 2: Upload Excel
# ======================
elif st.session_state.step == 'upload':
    st.write("### 🧾 Langkah 2 — Masukan file excel template")
    uploaded_file = st.file_uploader("📎 Upload file Excel", type="xlsx")
    type_spt = st.session_state.type_spt

    sheet_name = "DATA"      # Replace with your sheet name
    TIN_cell = "C1"
    start_row = 4

    # XML field tags in order for columns B → P
    xml_tags_21 = [
        "TaxPeriodMonth", "TaxPeriodYear", "CounterpartTin",
        "IDPlaceOfBusinessActivityOfIncomeRecipient", "StatusTaxExemption",
        "TaxCertificate", "TaxObjectCode", "Gross", "Deemed",
        "Rate", "Document", "DocumentNumber",
        "DocumentDate", "IDPlaceOfBusinessActivity",
        "WithholdingDate"
    ]

    xml_tags_uni = [
        "TaxPeriodMonth", "TaxPeriodYear", "CounterpartTin",
        "IDPlaceOfBusinessActivityOfIncomeRecipient", 
        "TaxCertificate", "TaxObjectCode", "TaxBase", 
        "Rate", "Document", "DocumentNumber",
        "DocumentDate", "IDPlaceOfBusinessActivity",
        "GovTreasurerOpt", "SP2DNumber",
        "WithholdingDate"
    ]

    type_spt = "Bp21" or "Bpu"
    if type_spt == "Bp21":
        xml_tags = xml_tags_21 
    elif type_spt == "Bpu":
        xml_tags = xml_tags_uni

    # Columns B → P in Excel
    excel_cols = list("BCDEFGHIJKLMNOP")

    # Columns for special evaluation
    evaluate_cols = ["E"]

    # --- SIMPLE FORMULA EVALUATOR ---
    def eval_formula(cell, ws):
        """
        Evaluates simple Excel formulas with:
        - Concatenation (&)
        - Basic arithmetic (+ - * /)
        - References like D4, F12 etc.
        """
        f = cell.value
        if not isinstance(f, str) or not f.startswith("="):
            return f
        expr = f[1:]  # remove leading "="
        
        # Replace cell references (e.g. D4) with actual values
        def repl_ref(match):
            ref = match.group(0)
            val = ws[ref].value
            # if referenced cell also has a formula, don't re-evaluate recursively here
            return f"'{val}'" if val is not None else "''"

        # Convert Excel & into Python +
        expr = re.sub(r"([A-Za-z]+[0-9]+)", repl_ref, expr)
        expr = expr.replace("&", "+")

        try:
            # Evaluate safely
            return str(eval(expr))
        except Exception as e:
            return ""
    
    if uploaded_file:
        excel_file = uploaded_file
        filename = uploaded_file.name.rsplit('.', 1)[0]
        # --- READ EXCEL ---
        wb = openpyxl.load_workbook(excel_file, data_only=False)
        ws = wb[sheet_name]

        # Get TIN
        TIN = str(ws[TIN_cell].value)

        # --- BUILD XML ---
        root = ET.Element(f"{type_spt}Bulk", {"xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance"})
        ET.SubElement(root, "TIN").text = TIN
        list_elem = ET.SubElement(root, f"ListOf{type_spt}")

        row = start_row
        while True:
            if ws[f"B{row}"].value is None:
                break

            bp_elem = ET.SubElement(list_elem, type_spt)

            for col_letter, xml_tag in zip(excel_cols, xml_tags):
                cell = ws[f"{col_letter}{row}"]

                # If this column needs evaluation AND has a formula:
                if col_letter in evaluate_cols and isinstance(cell.value, str) and cell.value.startswith("="):
                    cell_val = eval_formula(cell, ws)
                else:
                    # For normal, use data_only=True value if available
                    if cell.value is None and cell.data_type == "f":
                        # fallback if openpyxl stored a formula but no computed value
                        cell_val = eval_formula(cell, ws)
                    else:
                        cell_val = cell.value

                # Convert None to empty string
                text_val = "" if cell_val is None else str(cell_val)

                ET.SubElement(bp_elem, xml_tag).text = text_val

            row += 1

        # --- WRITE XML FILE ---
        tree = ET.ElementTree(root)
        output = BytesIO()
        tree.write(output, encoding="utf-8", xml_declaration=True)
        output.seek(0)

        st.download_button(
            label="Download XML File",
            data=output,
            file_name=f"{filename}.xml",
            mime=""
        )

    if st.button("⬅️ Kembali"):    
        go_to_step("pilihSPT")
        st.rerun()  
