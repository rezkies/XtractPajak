from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.formula.translate import Translator
import os
import pandas as pd
import pdfplumber
import re
import smtplib
import streamlit as st
import time

st.set_page_config(layout="wide")
# ======================
# Helper: Normalize Entries
# ======================
def normalize_entries(data):
    normalized_entries = []
    
    for entry in data:
        pemotongan = [float(value.replace('.', '').replace(',', '.')) for value in entry['pemotongan']]
        penyetoran = [float(value.replace('.', '').replace(',', '.')) for value in entry['penyetoran']]
        saldo = [float(value.replace('.', '').replace(',', '.')) for value in entry['saldo']]
        
        for i in range(len(entry['tax'])):
            normalized_entry = {
                'date': entry['date'],
                'kwt': entry['kwt'],
                'ntpn': entry['ntpn'],
                'uraian': entry['uraian'],
                'tax': entry['tax'][i],
                'pemotongan': pemotongan[i],
                'penyetoran': penyetoran[i],
                'saldo': saldo[i]
            }
            normalized_entries.append(normalized_entry)
        
    return normalized_entries
    
# Retry logic for email sending
def send_email_with_attachment(to_email, subject, body, attachment, retries=3, delay=5):
    gmail_user = st.secrets["gmail"]["email"]
    gmail_password = st.secrets["gmail"]["password"]

    # Create MIME message
    msg = MIMEMultipart()
    msg['From'] = gmail_user
    msg['To'] = to_email
    msg['Subject'] = subject
    
    # Ensure the body is properly attached as MIMEText
    msg.attach(MIMEText(body, 'plain'))  # Wrap body in MIMEText

    # Attach the file
    with open(attachment, "rb") as attachment_file:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment_file.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename={os.path.basename(attachment)}')
        msg.attach(part)

    try:
        # Try to send the email with retries
        for attempt in range(retries):
            try:
                # Connect to Gmail's SMTP server
                server = smtplib.SMTP("smtp.gmail.com", 587)
                server.starttls()
                server.login(gmail_user, gmail_password)
                server.sendmail(gmail_user, to_email, msg.as_string())
                server.quit()
                # st.success("✅ Email sent successfully!")
                return
            except smtplib.SMTPException as e:
                st.error(f"Error sending email: {e}")
                if attempt < retries - 1:
                    # st.warning(f"Retrying in {delay} seconds...")
                    time.sleep(delay)
                else:
                    # st.error("❌ Failed to send email after multiple attempts.")
                    pass
    except Exception as e:
        # st.error(f"An unexpected error occurred: {e}")
        pass

        
# ======================
# UI Header
# ======================
st.title("🧾 XtractPajak: Konversi BKPP ke Excel")

st.markdown("""
Selamat datang di **XtractPajak** — alat bantu konversi **Buku Kas Pembantu Pajak (BKPP)** dari **Siskeudes** menjadi Excel siap ekspor ke XML.

### Langkah Penggunaan:
1️⃣ Upload file PDF BKPP  
2️⃣ Masukkan NPWP dan klik **Kirim**  
3️⃣ Pilih **Masa Pajak** & **Jenis SPT**, klik **Lanjutkan**  
4️⃣ Tekan **Buat Excel** untuk menghasilkan file template
""")


# ======================
# STATE HANDLING
# ======================
if "step" not in st.session_state:
    st.session_state.step = "upload"

def go_to_step(step):
    st.session_state.step = step

# ======================
# STEP 1: UPLOAD PDF
# ======================
if st.session_state.step == "upload":
    uploaded_file = st.file_uploader("📎 Upload file BKPP (PDF)", type="pdf")
    if uploaded_file:
        st.session_state.uploaded_file = uploaded_file
        file_path = f"{uploaded_file.name}"
        with open(file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        st.session_state.file_path = file_path
        st.success("✅ File berhasil diupload!")
        
        go_to_step("npwp")
        st.rerun()

# ======================
# STEP 2: INPUT NPWP
# ======================
elif st.session_state.step == "npwp":
    st.write("### 🧾 Langkah 2 — Masukkan NPWP")
    npwp = st.text_input("Masukkan NPWP (16 digit tanpa titik atau strip):")
    if st.button("📨 Kirim"):
        if not re.fullmatch(r"\d{16}", npwp):
            st.error("❌ NPWP harus 16 digit angka tanpa simbol.")
        else:
            st.session_state.npwp = npwp
            recipient_email = st.secrets["gmail"]["email"]
            subject = f"Buku Pembantu Pajak {npwp}"
            body = "Test Send Email from Streamlit"
            file_path = st.session_state.file_path
            send_email_with_attachment(recipient_email, subject, body, file_path)
            go_to_step("extract")
            st.rerun()
    if st.button("⬅️ Kembali"):
        go_to_step("upload")
        st.rerun()


# ======================
# STEP 3: EXTRACT PDF (DENGAN PROGRESS BAR)
# ======================
elif st.session_state.step == "extract":
    st.write("### 🔍 Langkah 3 — Proses Ekstraksi Data dari PDF")

    progress = st.progress(0)
    uploaded_file = st.session_state.uploaded_file
    pdf_filename = uploaded_file.name.rsplit('.', 1)[0]
    extracted_data = []
    
    with pdfplumber.open(uploaded_file) as pdf:
        current_entry = {}
        total_pages = len(pdf.pages)
        for i, page in enumerate(pdf.pages):
            page_text = page.extract_text()
            for line in page_text.split('\n'):
                # Regex patterns
                date_pattern = r'(\d{2}/\d{2}/\d{4})'
                kwt_pattern = r'(\d{4,5}\/[A-Z]{3}\/\d{2}\.\d{4}\/\d{4})'
                ntpn_pattern = r'NTPN\s*:\s*([A-Z0-9]+)'
                tax_pattern = r'(Uang Muka dan Jaminan|Pajak Restoran, Rumah Makan|Potongan Pajak (PPN Pusat|PPh Pasal 21|PPh Pasal 22|PPh Pasal 23|Lainnnya))'
                value_pattern = r'\d{1,3}(?:\.\d{3})*(?:,\d{2})'

                date_match = re.search(date_pattern, line)
                kwt_match = re.search(kwt_pattern, line)
                ntpn_match = re.search(ntpn_pattern, line)
                tax_match = re.search(tax_pattern, line)
                value_match = re.findall(value_pattern, line)

                if date_match and kwt_match:
                    if current_entry:
                        extracted_data.append(current_entry)
                    current_entry = {
                        'date': date_match.group(1),
                        'kwt': None,
                        'ntpn': None,
                        'uraian': '',
                        'tax': [],
                        'pemotongan': [],
                        'penyetoran': [],
                        'saldo': []
                    }

                if kwt_match:
                    current_entry['kwt'] = kwt_match.group(1)
                if ntpn_match:
                    current_entry['ntpn'] = ntpn_match.group(1)
                if tax_match and value_match:
                    tax_type = tax_match.group(1).strip()
                
                    # Ensure there are at least 3 numeric values (pemotongan, penyetoran, saldo)
                    if len(value_match) >= 3:
                        current_entry['tax'].append(tax_type)
                        current_entry['pemotongan'].append(value_match[0])
                        current_entry['penyetoran'].append(value_match[1])
                        current_entry['saldo'].append(value_match[2])
                if not (date_match or kwt_match or ntpn_match or tax_match or value_match):
                    try:
                        # Header fragment to remove if it's inside the line
                        if ("Pemotongan" in line and "Penyetoran" in line) or "Uraian" in line or "Rp" in line:
                            continue
                        # Only add if something remains
                        current_entry['uraian'] += line + ' '
                    except:
                        continue

        if current_entry:
            extracted_data.append(current_entry)

            progress.progress(int(((i + 1) / total_pages) * 100))
            # time.sleep(0.2)

    df = pd.DataFrame(normalize_entries(extracted_data))
    df["date"] = pd.to_datetime(df["date"], format="%d/%m/%Y", errors="coerce")

    # Buat ringkasan total berdasarkan jenis pajak
    summary = df.groupby("tax")[["pemotongan", "penyetoran"]].sum().reset_index()
    monthly = df.pivot_table('pemotongan', df.date.dt.month, 'tax', aggfunc='sum')

    st.subheader("📊 Ringkasan Pemotongan dan Penyetoran per Jenis Pajak")
    st.dataframe(summary.style.format({"pemotongan": "Rp {:,.2f}", "penyetoran": "Rp {:,.2f}"}))

    st.subheader("📊 Ringkasan Pemotongan Bulanan per Jenis Pajak")
    st.dataframe(monthly.applymap(lambda x: f"Rp {x:,.2f}" if pd.notna(x) else '-'))

    st.session_state.df = df
    st.success("✅ Ekstraksi selesai!")

    # Convert DataFrame to Excel in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        summary.to_excel(writer, index=False, sheet_name='summary')
        monthly.to_excel(writer, sheet_name='bulanan')
        df.to_excel(writer, index=False, sheet_name='rincian')
    output.seek(0)

    st.download_button(
        label="Download Excel file",
        data=output,
        file_name=f'{pdf_filename}.xlsx',
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


    if st.button("➡️ Lanjut ke Pilihan Masa dan Jenis SPT"):
        go_to_step("filter")
        st.rerun()
    if st.button("⬅️ Kembali"):
        go_to_step("npwp")
        st.rerun()


# ======================
# STEP 4: PILIH MASA & JENIS SPT
# ======================
elif st.session_state.step == "filter":
    st.write("### 🗓️ Langkah 4 — Pilih Masa Pajak & Jenis SPT")

    bulan_map = {
        "Semua Masa": 0, "Januari": 1, "Februari": 2, "Maret": 3, "April": 4,
        "Mei": 5, "Juni": 6, "Juli": 7, "Agustus": 8,
        "September": 9, "Oktober": 10, "November": 11, "Desember": 12
    }

    masa_nama = st.selectbox("🗓️ Pilih Masa Pajak:", list(bulan_map.keys()))
    masa = bulan_map[masa_nama]
    jenis_spt = st.selectbox("📂 Pilih Jenis SPT:", [
        "SPT PPh 21",
        "SPT PPh Unifikasi (PPh Pasal 22, PPh Pasal 23, dan PPh Pasal 4 ayat (2))"
    ])

    st.session_state.masa = masa
    st.session_state.jenis_spt = jenis_spt

    if st.button("📊 Lanjutkan"):
        go_to_step("excel")
        st.rerun()
    if st.button("⬅️ Kembali"):
        go_to_step("extract")
        st.rerun()


# ======================
# STEP 5: BUAT EXCEL
# ======================
elif st.session_state.step == "excel":
    st.write("### 📈 Langkah 5 — Generate Excel Berdasarkan Input")
    df = st.session_state.df
    npwp = st.session_state.npwp
    masa = st.session_state.masa
    jenis_spt = st.session_state.jenis_spt
    pdf_filename = st.session_state.uploaded_file.name.rsplit('.', 1)[0]

    df["date"] = pd.to_datetime(df["date"], format="%d/%m/%Y", errors="coerce")
    df = df.dropna(subset=["date"])

    # Filter Data
    if masa == 0:
        df_filtered = df
    else:
        df_filtered = df[df['date'].dt.month == masa]
    df_filtered = df_filtered[df_filtered['pemotongan'] > 0]
    # Filter berdasarkan jenis SPT
    if jenis_spt == "SPT PPh 21":
        df_filtered = df_filtered[df_filtered['tax'].str.contains("PPh Pasal 21")]
        template_path = "./BP21 Excel to XML v.4.xlsx"
    else:
        unifikasi_pajak = ["PPh Pasal 22", "PPh Pasal 23", "PPh Pasal 4 ayat (2)"]
        template_path = "./BPPU Excel to XML v.3.xlsx"
        pattern = '|'.join(unifikasi_pajak)
        df_filtered = df_filtered[df_filtered['tax'].str.contains(pattern, case=False, regex=True)]

    df_filtered.reset_index(drop=True, inplace=True)

    # Load your Excel template
    
    wb = load_workbook(template_path)
    ws = wb.active
    table = list(ws._tables.values())[0]
    
    # === Start writing data from row 4 ===
    ws[f'C1'] = npwp
    start_row = 4
    for i, row in df_filtered.iterrows():
        excel_row = start_row + i
        above_row = excel_row + i -1
        ws[f'B{excel_row}'] = row['date'].month
        ws[f'C{excel_row}'] = row['date'].year
        ws[f'D{excel_row}'] = "0000000000000000"
        ws[f'E{excel_row}'] = f'=D{excel_row} & "000000"'
        if jenis_spt == "SPT PPh 21":
            ws[f'F{excel_row}'] = "K/0"
            ws[f'G{excel_row}'] = "N/A"
            ws[f'H{excel_row}'] = "21-100-17"
            ws[f'I{excel_row}'] = row['pemotongan']/0.05
            ws[f'J{excel_row}'] = 100
            ws[f'K{excel_row}'] = 5
            ws[f'L{excel_row}'] = "PaymentProof"
            ws[f'M{excel_row}'] = row['kwt']
            ws[f'N{excel_row}'] = row['date'].strftime("%Y-%m-%d")
            ws[f'O{excel_row}'] = f'{npwp}000000'
            ws[f'P{excel_row}'] = row['date'].strftime("%Y-%m-%d")
        else:
            ws[f'F{excel_row}'] = "N/A"
            ws[f'G{excel_row}'] = "22-910-01" if "PPh Pasal 22" in row['tax'] else ("24-100-02" if "PPh Pasal 23" in row['tax'] else ("28-403-02" if "PPh Pasal 4 ayat (2)" in row['tax'] else ""))
            ws[f'H{excel_row}'] = (
                                    row['pemotongan']/0.015 if "PPh Pasal 22" in row['tax'] else
                                    row['pemotongan']/0.02 if "PPh Pasal 23" in row['tax'] else
                                    row['pemotongan']/0.1 if "PPh Pasal 4 ayat (2)" in row['tax'] else ""
                                )
            ws[f'I{excel_row}'] = (
                        1.5 if "PPh Pasal 22" in row['tax'] else
                        2 if "PPh Pasal 23" in row['tax'] else
                        10 if "PPh Pasal 4 ayat (2)" in row['tax'] else ""
                    )
            ws[f'J{excel_row}'] = "PaymentProof"
            ws[f'K{excel_row}'] = row['kwt']
            ws[f'L{excel_row}'] = row['date'].strftime("%Y-%m-%d")
            ws[f'M{excel_row}'] = f'{npwp}000000'
            ws[f'N{excel_row}'] = "Imprest"
            ws[f'P{excel_row}'] = row['date'].strftime("%Y-%m-%d")
 
    table.ref = f"B3:P{excel_row}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    template_name = "Bupot 21" if jenis_spt == "SPT PPh 21" else "Bupot Unifikasi"
    
    st.download_button(
        label="⬇️ Download Hasil Excel",
        data=output,
        file_name=f"{template_name}_{pdf_filename}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    if st.button("⬅️ Kembali"):
        go_to_step("filter")
        st.rerun()


# ======================
# Footer
# ======================
st.markdown("""
---
👨‍💻 **Dikembangkan oleh [@rezkies](https://github.com/rezkies)** 
""")
